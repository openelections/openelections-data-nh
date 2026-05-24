"""Read .xls and .xlsx workbooks behind one interface.

We dispatch on file content (magic bytes) rather than trusting the extension —
the NH SoS site has been known to label files inconsistently, and the cost of
sniffing four bytes is negligible.
"""

from __future__ import annotations

import pathlib
from typing import Any, Protocol


XLS_MAGIC = b"\xd0\xcf\x11\xe0"  # OLE2 compound document (legacy .xls)
XLSX_MAGIC = b"PK\x03\x04"        # ZIP archive (OOXML .xlsx)


class _SheetLike(Protocol):
    nrows: int
    ncols: int

    def cell_value(self, row: int, col: int) -> Any: ...
    def row_values(self, row: int) -> list[Any]: ...


class WorkbookReader:
    """A small, format-agnostic view over one sheet of a workbook.

    Sheet selection is by index (default 0). The interface deliberately mirrors
    xlrd so existing code patterns translate without much thought.

    For multi-sheet workbooks, callers should iterate over `sheet_names(path)`
    or `sheet_count(path)` and construct one `WorkbookReader` per sheet.
    """

    def __init__(self, path: pathlib.Path, sheet_index: int = 0):
        self._path = pathlib.Path(path)
        self._sheet_index = sheet_index
        self._sheet: _SheetLike = self._open()
        self._sheet_name: str = self._open_sheet_name()

    @property
    def sheet_name(self) -> str:
        """Name of the sheet this reader is bound to."""
        return self._sheet_name

    @classmethod
    def sheet_names(cls, path: pathlib.Path) -> list[str]:
        """List all sheet names in the workbook, in order."""
        fmt = cls.sniff_format(path)
        if fmt == "xls":
            import xlrd
            wb = xlrd.open_workbook(str(path), on_demand=True)
            try:
                return list(wb.sheet_names())
            finally:
                wb.release_resources()
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    @classmethod
    def sheet_count(cls, path: pathlib.Path) -> int:
        """Number of sheets in the workbook."""
        return len(cls.sheet_names(path))

    @classmethod
    def sniff_format(cls, path: pathlib.Path) -> str:
        """Return 'xls' or 'xlsx' based on the first few bytes of the file."""
        with open(path, "rb") as fh:
            head = fh.read(8)
        if head.startswith(XLS_MAGIC):
            return "xls"
        if head.startswith(XLSX_MAGIC):
            return "xlsx"
        raise ValueError(
            f"{path}: not recognized as .xls (OLE2) or .xlsx (OOXML); "
            f"first 8 bytes = {head!r}"
        )

    def _open(self) -> _SheetLike:
        fmt = self.sniff_format(self._path)
        if fmt == "xls":
            return _XlrdSheet(self._path, self._sheet_index)
        return _OpenpyxlSheet(self._path, self._sheet_index)

    def _open_sheet_name(self) -> str:
        """Best-effort retrieval of the sheet's name."""
        names = self.sheet_names(self._path)
        if 0 <= self._sheet_index < len(names):
            return names[self._sheet_index]
        return ""

    @property
    def nrows(self) -> int:
        return self._sheet.nrows

    @property
    def ncols(self) -> int:
        return self._sheet.ncols

    def cell_value(self, row: int, col: int) -> Any:
        return self._sheet.cell_value(row, col)

    def row_values(self, row: int) -> list[Any]:
        return self._sheet.row_values(row)


class _XlrdSheet:
    def __init__(self, path: pathlib.Path, sheet_index: int):
        import xlrd
        wb = xlrd.open_workbook(str(path))
        self._sheet = wb.sheet_by_index(sheet_index)
        self.nrows = self._sheet.nrows
        self.ncols = self._sheet.ncols

    def cell_value(self, row: int, col: int) -> Any:
        return self._sheet.cell_value(row, col)

    def row_values(self, row: int) -> list[Any]:
        return self._sheet.row_values(row)


class _OpenpyxlSheet:
    def __init__(self, path: pathlib.Path, sheet_index: int):
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        names = wb.sheetnames
        self._sheet = wb[names[sheet_index]]
        # openpyxl is 1-indexed and reports max_row/max_col; normalize to xlrd-style 0-indexed counts.
        self._rows: list[list[Any]] = []
        for row in self._sheet.iter_rows(values_only=True):
            self._rows.append([_normalize_cell(v) for v in row])
        self.nrows = len(self._rows)
        self.ncols = max((len(r) for r in self._rows), default=0)
        # Pad short rows so cell_value never IndexErrors within (nrows, ncols).
        for r in self._rows:
            if len(r) < self.ncols:
                r.extend([""] * (self.ncols - len(r)))

    def cell_value(self, row: int, col: int) -> Any:
        return self._rows[row][col]

    def row_values(self, row: int) -> list[Any]:
        return list(self._rows[row])


def _normalize_cell(v: Any) -> Any:
    """Coerce openpyxl's None to '' so it matches xlrd's empty-cell behavior."""
    return "" if v is None else v
